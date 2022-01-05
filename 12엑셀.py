import openpyxl

wb = openpyxl.Workbook() #워크북 약자로 동작입력

ws=wb.create_sheet('squidg')

ws['A1'] = '참가번'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

wb.save('참가자_data.xlsx')

