import openpyxl
fpath = r'C:\Pyth\참가자_data.xlsx'
wb = openpyxl.load_workbook(fpath)

ws = wb['squidg']

ws['A3'] = 456
ws['B3'] = '성'

wb.save(fpath)

